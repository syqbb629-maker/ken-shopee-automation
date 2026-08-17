"""
shopee_excel_generator.py
----------------------------
V2仕様19・20・21・27の実装。

- 独自Excel形式は作らず、Shopee公式Mass Uploadテンプレート
  (templates/shopee_sg_template.xlsx, shopee_my_template.xlsx)を
  コピーして商品入力行だけを更新する。
- 公式テンプレートが用意されていない場合、偽物のテンプレートを
  作って「公式」として扱わない。明確なエラーで案内する。
- READY_TO_LISTの商品のみを書き込む(REVIEW/LOW_PROFIT等は含めない)。

公式テンプレートの列名はShopeeの仕様変更やSG/MYの差異で変わりうるため、
本モジュールはテンプレートのヘッダー行(1行目、または指定行)を読み取り、
"列名 -> 書き込む値" のマッピング辞書(COLUMN_MAP)に基づいて
一致する列だけへ値を書き込む。テンプレートに存在しない列は無視する
(テンプレートの構造・隠しシート・データ検証を壊さないため)。
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

# Shopee公式テンプレートの列名 -> 商品マスター側で書き込む値を得るための項目名
# 実際の公式テンプレートの列名(英語表記の揺れ)に合わせて追記・修正してください。
COLUMN_MAP = {
    "Product Name": "product_name_en",
    "Product Name*": "product_name_en",
    "Product Description": "description_en",
    "Product Description*": "description_en",
    "Category": "category",
    "Category*": "category",
    "Brand": "brand",
    "Price": "_price",       # market別に置換
    "Price*": "_price",
    "Stock": "stock",
    "Stock*": "stock",
    "Parent SKU": "_sku",    # market別に置換
    "SKU": "_sku",
    "SKU*": "_sku",
    "Cover image": "image_main_url",
    "Cover Image": "image_main_url",
    "Cover Image*": "image_main_url",
    "Item Image 1": "image_main_url",
    "Item Image 2": "image_2_url",
    "Item Image 3": "image_3_url",
    "Weight": "shipping_weight_g",
    "Weight*": "shipping_weight_g",
    "Weight (kg)": "shipping_weight_kg",
    "Weight (kg)*": "shipping_weight_kg",
    "Length": "package_length_cm",
    "Length(cm)": "package_length_cm",
    "Width": "package_width_cm",
    "Width(cm)": "package_width_cm",
    "Height": "package_height_cm",
    "Height(cm)": "package_height_cm",
    "Shipping Channel": "_shipping_channel",
    "Pre-order DTS": "_dts",
}

DEFAULT_SHIPPING_CHANNEL = "Standard"
DEFAULT_DTS = "2"


class TemplateNotFoundError(FileNotFoundError):
    pass


def _find_header_row(ws, max_scan_rows: int = 10) -> int:
    """Shopee公式テンプレートは説明行・隠し行を含むことが多いため、
    「Product Name」のような既知列名が最も多く見つかる行をヘッダー行とみなす。
    """
    best_row = 1
    best_score = -1
    for r in range(1, max_scan_rows + 1):
        values = [str(c.value).strip() if c.value is not None else "" for c in ws[r]]
        score = sum(1 for v in values if v in COLUMN_MAP)
        if score > best_score:
            best_score = score
            best_row = r
    return best_row


def _row_to_dict(row: pd.Series, market: str) -> dict:
    d = row.to_dict()
    if market == "SG":
        d["_price"] = row.get("sg_price", "")
        d["_sku"] = row.get("sku_sg", "")
    else:
        d["_price"] = row.get("my_price", "")
        d["_sku"] = row.get("sku_my", "")
    try:
        d["shipping_weight_kg"] = float(row.get("shipping_weight_g", 0) or 0) / 1000.0
    except (ValueError, TypeError):
        d["shipping_weight_kg"] = ""
    d["_shipping_channel"] = DEFAULT_SHIPPING_CHANNEL
    d["_dts"] = DEFAULT_DTS
    return d


def generate_shopee_excel(
    ready_df: pd.DataFrame,
    template_path: Path,
    output_path: Path,
    market: str,
) -> int:
    """READY_TO_LISTの商品だけをShopee公式テンプレートに書き込み保存する。
    戻り値: 書き込んだ行数
    """
    if not template_path.exists():
        raise TemplateNotFoundError(
            f"Shopee公式テンプレートが見つかりません: {template_path}\n"
            "本物のShopee公式Mass Uploadテンプレート(.xlsx)をこのパスに配置してください。\n"
            "偽のテンプレートで代用することはできません(README参照)。"
        )

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    header_row_idx = _find_header_row(ws)
    header_cells = ws[header_row_idx]
    col_index_by_name = {}
    for cell in header_cells:
        name = str(cell.value).strip() if cell.value is not None else ""
        if name:
            col_index_by_name[name] = cell.column

    write_row = header_row_idx + 1
    written = 0

    for _, row in ready_df.iterrows():
        data = _row_to_dict(row, market)
        for template_col_name, source_field in COLUMN_MAP.items():
            if template_col_name not in col_index_by_name:
                continue
            col_idx = col_index_by_name[template_col_name]
            value = data.get(source_field, "")
            ws.cell(row=write_row, column=col_idx, value=value)
        write_row += 1
        written += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return written
