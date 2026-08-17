from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from src.shopee_excel_generator import generate_shopee_excel, TemplateNotFoundError

FIXTURE_DIR = Path("tests/fixtures")
FIXTURE_TEMPLATE = FIXTURE_DIR / "fake_shopee_sg_template.xlsx"


def _make_fixture_template():
    FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "Product Name*", "Product Description*", "Category*", "Brand",
        "Price*", "Stock*", "SKU*", "Cover Image*",
        "Weight*", "Length(cm)", "Width(cm)", "Height(cm)",
    ]
    ws.append(headers)
    wb.save(FIXTURE_TEMPLATE)


def _sample_df(status: str) -> pd.DataFrame:
    return pd.DataFrame([
        {
            "item_id": "T001",
            "product_name_en": "Test Product 40g",
            "description_en": "Brand: Test\nShips from Japan",
            "category": "Skincare",
            "brand": "TestBrand",
            "sg_price": "32.9",
            "my_price": "99.9",
            "stock": "2",
            "sku_sg": "TB-SG-001",
            "sku_my": "TB-MY-001",
            "image_main_url": "https://example.com/img.jpg",
            "image_2_url": "",
            "image_3_url": "",
            "shipping_weight_g": "120",
            "package_length_cm": "4",
            "package_width_cm": "4",
            "package_height_cm": "12",
            "listing_status": status,
        }
    ])


def test_missing_template_raises_clear_error(tmp_path):
    df = _sample_df("READY_TO_LIST")
    missing_template = tmp_path / "does_not_exist.xlsx"
    output_path = tmp_path / "out.xlsx"
    with pytest.raises(TemplateNotFoundError):
        generate_shopee_excel(df, missing_template, output_path, market="SG")


def test_ready_rows_written_to_official_template(tmp_path):
    _make_fixture_template()
    df = _sample_df("READY_TO_LIST")
    output_path = tmp_path / "shopee_sg_ready.xlsx"

    written = generate_shopee_excel(df, FIXTURE_TEMPLATE, output_path, market="SG")

    assert written == 1
    assert output_path.exists()

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    data_row = [c.value for c in ws[2]]
    row_dict = dict(zip(header, data_row))
    assert row_dict["Product Name*"] == "Test Product 40g"
    assert row_dict["Price*"] == "32.9"
    assert row_dict["SKU*"] == "TB-SG-001"


def test_non_ready_rows_are_not_passed_in(tmp_path):
    """呼び出し側がREADY_TO_LISTのみを渡す前提を確認する
    (ready_dfへのフィルタはpipeline側の責務)。"""
    _make_fixture_template()
    df = _sample_df("READY_TO_LIST")
    non_ready_df = df[df["listing_status"] == "LOW_PROFIT"]  # 空になるはず
    output_path = tmp_path / "shopee_sg_ready.xlsx"

    written = generate_shopee_excel(non_ready_df, FIXTURE_TEMPLATE, output_path, market="SG")
    assert written == 0
